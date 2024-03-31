import openpyxl
import xlwings as xw
import csv

#region Const Values
positionColumn = 1
teamColumn = 2
playerColumn = 3
positionMainValueColumn = 10
positionSecondaryValueColumn = 12
positionProgressColumn = 14
CAValueColumn = 21
CAProgressColumn = 22
allColumnsCount = 27
csvTemplateFile = "template.csv"
valuesColumnsDictionary = {
    "Position": 0,
    "Team": 1,
    "Name": 2,
    "Age": 3,
    "Country": 4,
    "ToolRating_1": 6,
    "ToolRating_2": 7,
    "PositionMain_Name": 9,
    "PositionMain_Value": 10,
    "PositionSec_Name": 11,
    "PositionSec_Value": 12,
    "Progress": 14,
    "Determination": 17,
    "Potential": 18,
    "NoPermission": 19,
    "CA": 21,   
    "PA": 23,
    "PlayerStatus": 24,
    "RaportStatus": 25,
    "Info": 26
}
templateRow = ['BR', 'REZERWA', 'EmptyName', 0, 'ENG', None, 0, 0, None, 'BR ', 0, 'BR-Lib', 0, None, 'NEW', None, None, None, None, None, None, 0, "NEW", 0, None, None, None]
#endregion

#TODO column by index, or finding index by column name? 
#TODO use only one: xlwings or openpyxl?

class Helper:
    #only for openpyxl
    def find_row_by_value(self, sheet, column, min_row_value, value):        
        for index, row in enumerate(sheet.iter_rows(min_row = min_row_value, max_row=sheet.max_row, values_only=True)):
            if row[column-1] == value:
                return index        
        return None            

    #only for xlwings
    def remove_row(self, sheet, row):
        sheet.range((row, 1)).api.EntireRow.Delete()
        
class OpenpyxlController:
    def __init__(self, excelFileName):
        self.excelFileName = excelFileName

    def create_sheet(self):
        self.wb = openpyxl.load_workbook(self.excelFileName)
        return self.wb.active

    def close_controller(self):
        self.wb.close()    

class XlwingsController:
    def __init__(self, excelFileName):
        self.excelFileName = excelFileName

    def create_sheet(self, useApp):
        self.useApp = useApp
        if useApp:            
            self.app = xw.App(visible=False) #use it for "silent" open & close excel
        self.wb = xw.Book(self.excelFileName)
        return self.wb.sheets.active
    
    def close_controller(self, closeFile):
        self.wb.save()
        if(closeFile):
            self.wb.close()
        if(self.useApp):
            self.app.quit()

class CsvController:
    def create_csv(self, filename, data):
        with open(filename, 'w', newline='') as csvfile:
            writer = csv.writer(csvfile)
            for row in data:
                writer.writerow(row)

    def read_csv(self, filename):
        data = []
        with open(filename, 'r', newline='') as csvfile:
            reader = csv.reader(csvfile)
            for row in reader:
                data.append(row)
        return data
    
    def createTemplateFile(self):
        valuesTemplateFilename = csvTemplateFile
        data = [
            ['Position', 'Team', 'Name', 'Age', 'Country', 'NONE', 'ToolRating_1', 'ToolRating_2', 'NONE', 'PositionMain_Name', 'PositionMain_Value', 'PositionSec_Name', 'PositionSec_Value', 'NONE', 
            'Progress', 'NONE', 'NONE', 'Determination', 'Potential', 'NoPermission', 'NONE', 'CA', 'CAChange', 'PA', 'PlayerStatus', 'RaportStatus', 'Info']    
        ]
        self.create_csv(valuesTemplateFilename, data)

class ExcelModificationsController:
    def __init__(self, excelFileName): 
        self.helper = Helper()
        self.openpyxlController = OpenpyxlController(excelFileName)
        self.xlwingsController = XlwingsController(excelFileName)    
        self.csvController = CsvController()            
        self.csvController.createTemplateFile()

    def get_teamPart_row(self, positionValue, teamValue):
        sheet = self.openpyxlController.create_sheet()
        positionRow = self.helper.find_row_by_value(sheet, positionColumn, 1, positionValue)
        if positionRow is not None:        
            team_part_row = self.helper.find_row_by_value(sheet, teamColumn, positionRow, teamValue)
            self.openpyxlController.close_controller()
            if team_part_row is not None:
                return team_part_row + positionRow
            else:
                return None                    
        
    def insert_row(self, positionValue, teamValue):        
        sheet = self.xlwingsController.create_sheet(True)    
        teamPartRow = self.get_teamPart_row(positionValue, teamValue)       
        sheet.range((teamPartRow + 1, 1)).api.EntireRow.Insert()
        self.xlwingsController.close_controller(True)
        return teamPartRow
    
    def update_row_values(self, row_number, values):        
        sheet = self.xlwingsController.create_sheet(True)         
        sheet.range((row_number, 1), (row_number, len(values))).value = values
        self.xlwingsController.close_controller(True)

    def get_player_row(self, value):        
        sheet = self.openpyxlController.create_sheet()                
        playerRow = self.helper.find_row_by_value(sheet, playerColumn, 0, value)
        self.openpyxlController.close_controller()
        return playerRow
    
    def get_player_data_by_value(self, value):
        sheet = self.xlwingsController.create_sheet(True)
        rowNumber = self.get_player_row(value) + 1              
        data = sheet.range((rowNumber, 1), (rowNumber, allColumnsCount)).value
        self.xlwingsController.close_controller(True)
        return data

    def get_player_data_by_row(self, rowNumber):
        sheet = self.xlwingsController.create_sheet(True)
        data = sheet.range((rowNumber + 1, 1), (rowNumber + 1, allColumnsCount)).value
        self.xlwingsController.close_controller(True)
        return data
    
    def update_player_by_file(self, csvFileName):
        readPlayerData = self.csvController.read_csv(csvFileName)        
        nameValueColumn = readPlayerData[0].index("Name")        

        for i in range(len(readPlayerData) - 1):
            playerRow = self.get_player_row(readPlayerData[i+1][nameValueColumn])            
            if playerRow is None:
                newPlayerRow = self.insert_row(readPlayerData[i + 1][readPlayerData[0].index("Position")], readPlayerData[i + 1][readPlayerData[0].index("Team")])
                playerData = templateRow
                playerData = self.update_values(readPlayerData, playerData, i)                    
                self.update_row_values(newPlayerRow + 1, playerData)
            else:
                playerData = self.get_player_data_by_row(playerRow) 

                playerData[positionProgressColumn] = max(self.calculate_difference(playerData[positionMainValueColumn], readPlayerData[i + 1][readPlayerData[0].index("PositionMain_Value")]), self.calculate_difference(playerData[positionSecondaryValueColumn], readPlayerData[i + 1][readPlayerData[0].index("PositionSec_Value")]))         
                playerData[CAProgressColumn] = self.calculate_difference(playerData[CAValueColumn], readPlayerData[i + 1][readPlayerData[0].index("CA")])

                playerData = self.update_values(readPlayerData, playerData, i)                                          
                self.update_row_values(playerRow + 1, playerData)

    def update_values(self, readPlayerData, playerData, index):
        for valueName in readPlayerData[0]:
            if valueName in valuesColumnsDictionary:                
                playerData[valuesColumnsDictionary[valueName]] = readPlayerData[index + 1][readPlayerData[0].index(valueName)]                                 
        return playerData

    def delete_player_by_file(self, csvFileName): #TODO
        readPlayerData = self.csvController.read_csv(csvFileName)        
        playerRow = self.get_player_row(readPlayerData[1][readPlayerData[0].index("Name")])
        playerData = self.get_player_data_by_row(playerRow)
        if self.get_player_data_by_row(playerRow - 1)[1] != playerData[1] and self.get_player_data_by_row(playerRow + 1)[1] != playerData[1]:            
            sheet = self.xlwingsController.create_sheet(True)
            playerRow += 1            
            sheet.range((playerRow, 3), (playerRow, allColumnsCount)).value = None
        else:
            sheet = self.xlwingsController.create_sheet(True)
            playerRow += 1
            self.helper.remove_row(sheet, playerRow)
        self.xlwingsController.close_controller(True)
    
    def calculate_difference(self, previousValue, currentValue):
        previousValue = float(previousValue)
        currentValue = float(currentValue) 
        return (currentValue - previousValue)

excelModificationsController = ExcelModificationsController("Barrow.xlsx")
#excelModificationsController.insert_row("BR", "REZERWA")           
#excelModificationsController.update_player_by_file("playersFewInfo.csv")
#excelModificationsController.update_player_by_file("newPlayerOnlyFew.csv")
excelModificationsController.delete_player_by_file("playerToRemove.csv")